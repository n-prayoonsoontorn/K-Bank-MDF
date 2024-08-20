from collections import Counter
import json
import os
import openpyxl
import utility




def load_workbook(excel_file, sheet_name):
    try:
        workbook = openpyxl.load_workbook(excel_file,data_only=True)
        if sheet_name in workbook.sheetnames:
            selected_sheet = workbook[sheet_name]
            print(f"Loaded worksheet '{sheet_name}' from Excel file '{excel_file}' successfully.")
            print(f"Number of rows: {selected_sheet.max_row}, Number of columns: {selected_sheet.max_column}")

            utility.replace_stars(selected_sheet)
            return selected_sheet
        else:
            print(f"Worksheet '{sheet_name}' not found in the Excel file.")
            return None
    except Exception as e:
        print(f"Error loading the Excel file: {e}")
        return None   
  

def search_keywords(worksheet):

    keywords_right = {
        "Job Name": None,
        # "Source File Name/ File Path": None,
        "Target Schema Name": None,
        "Target Table Name": None,

    }
    keywords_below = {
        "Source System Name": None
    }
    keywords_right_list = {
        "Job Name": None,

    }

    ##-------------CUSTOM SEARCH--------- ##
    keyword_source_filename = {
        "Source File Name/ File Path": None,

    }

    results_source_filename  = utility.search_keywords_source_filename(worksheet,keyword_source_filename)



    results_right = utility.search_keywords_right(worksheet,keywords_right)
    results_right_list = utility.search_keywords_right_list(worksheet,keywords_right_list)

    results_below = utility.search_keywords_below(worksheet,keywords_below)
    
    results = {**results_right, **results_below,**results_right_list,**results_source_filename}
    return results

if __name__ == "__main__":
 


    path_in = "/Users/n.prayoonsoontorn/Downloads/script_gen_audit_file/input/"
    path_out = "/Users/n.prayoonsoontorn/Downloads/script_gen_audit_file/output/"


    sheet_name = "mdp_request_field"
    
    count = 0
    failed = []
    #---------  for investigation ---------
    parquet_file = []
    uniq_job_name = []
    text_file = []

    base_dir = []
    #--------------------------------------
    for dir_name in os.listdir(path_in):
        if dir_name != ".DS_Store":
            dir_path = os.path.join(path_in, dir_name)
            if os.path.isdir(dir_path):
                for filename in os.listdir(dir_path):
                    if filename != ".DS_Store":
                        count += 1
                        file_path = os.path.join(dir_path, filename)
                        worksheet = load_workbook(file_path, sheet_name)
                        results = search_keywords(worksheet)


                        #---------  Correction Field (format,etc.) ---------
                        job_name =results["Job Name"][0]
                        table_name = results["Target Table Name"]
                        schema_name = results["Target Schema Name"]
                        src_sys = results["Source System Name"]
                        fileNameOut = "audit_config_"+schema_name+"_"+table_name
                        path_root_context = "/Volumes/mdp{{env}}/artifact/data_quality/gx"
                        expectation_suite_name = "tech_vld_suite_"+schema_name+"_"+table_name
                        validator_name = schema_name+"_"+table_name+"_checkpoint"
                        


                        indexBody = utility.find_row_index_with_keywords_in_column_b_view_name(worksheet,"Body")
                        indexTailerLabel = utility.find_row_index_with_keywords_in_column_b_view_name(worksheet,"Tailor Label")

                        if indexBody is None or indexTailerLabel is None:
                            indexBody = utility.find_row_index_with_keywords_in_column_b_view_name(worksheet, "V Seq.") 
                            indexTailerLabel = worksheet.max_row+1   

                        columns_name_value = utility.get_column_values_in_range_(worksheet,"M",indexBody,indexTailerLabel)

                        columns_mandatory_value = utility.get_column_values_in_range_(worksheet,"N",indexBody,indexTailerLabel)
                        combined_list = list(zip(columns_name_value, columns_mandatory_value))
                        # print(combined_list)
                        combined_list = [item for item in combined_list if item[0] is not None and item[1] is not None]
                        ##remove "O" optional out 

                        print("---------------------------",combined_list)

                        filtered_list = [item for item in combined_list if 'M' in item[1]]
                        print("---------------------",filtered_list)
                        filtered_list_for_pk = [item for item in combined_list if 'PK' in item[1]]
                        filtered_column_name_pk = [item[0] for item in filtered_list_for_pk]
                        
                        filtered_column_names = [item[0] for item in filtered_list]

                       
                        print(filtered_list_for_pk)

                        filtered_column_name_pk = [item.strip() for item in filtered_column_name_pk if item is not None]

                        exclude_columns = ["pos_dt", 'load_tms', 'src_sys_id', 'ptn_yyyy', 'ptn_mm', 'ptn_dd', "upd_tms"]

                        expectations_list = []


                        if len(filtered_column_name_pk) > 1 :
                                expect_to_be_unique = {
                                "expectation_type": "expect_compound_columns_to_be_unique",
                                "kwargs": {
                                    "column_list": filtered_column_name_pk
                                },
                                "meta": {
                                    "level": "WARNING",
                                    "rule_id": "TECH_VLD_0001"
                                }
                            }
                                expectations_list.append(expect_to_be_unique)
                             
                                

                        if len(filtered_column_name_pk) == 1 :
                                expect_to_be_unique = {
                                "expectation_type": "expect_column_values_to_be_unique",
                                "kwargs": {
                                    "column": filtered_column_name_pk[0]
                                },
                                "meta": {
                                    "level": "WARNING",
                                    "rule_id": "TECH_VLD_0001"
                                }
                            }        
                                expectations_list.append(expect_to_be_unique)


                    ##apply rule to columns
                        for column_name, manda_value in filtered_list:                  
                            column_name = column_name.strip()
                            # if column_name in exclude_columns and column_name not in filtered_column_name_pk:

                            if column_name in exclude_columns:
                                    continue 
                            
                            expectation_not_be_null = {
                                "expectation_type": "expect_column_values_to_not_be_null",
                                "kwargs": {
                                    "column": column_name
                                },
                                "meta": {
                                    "level": "WARNING",
                                    "rule_id": "TECH_VLD_0002"
                                }
                            }
                            
                            # json_data["expectations"].append(expectation_not_be_null)
                            expectations_list.append(expectation_not_be_null)

 

                    # Define the JSON data for each job
                        json_data =  {
                                "data_source": {
                                    "asset": table_name, #TBC 
                                    "context_root_dir": path_root_context,
                                    "index_column_names": filtered_column_name_pk,
                                    "name": src_sys #Fix
                                },
                                "expectation_suites":[
                                    {
                                    "expectation_suite_name": expectation_suite_name,
                                    "expectations": expectations_list
                                },
                                ],
                                "validators":[
                                    {
                                    "expectation_suite_name": expectation_suite_name,
                                    "validator_name": validator_name
                                    }
                                ]
                            }
                        
                        

                        # Print or use the JSON object
                        json_string = json.dumps(json_data, indent=4)

                    
                        output_file = fileNameOut + ".json"
                        output_path = os.path.join(path_out, dir_name, output_file)
                        os.makedirs(os.path.dirname(output_path), exist_ok=True)


                        try:
                            with open(output_path, "w") as json_file:
                                print(output_path)
                                json.dump(json_data, json_file, indent=4, ensure_ascii=False)
                            print(f"\nJSON data has been written to {output_file}\n")
                        except Exception as e:
                            print(f"Failed to write JSON data to {output_file}: {e}")
                            failed.append(filename)
        
    print("Done writing", count, "files to", path_out)
    print("Failed files:", failed)









