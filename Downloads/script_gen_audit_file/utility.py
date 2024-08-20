import re
import os
import openpyxl


def get_column_values_in_range_(cbr_file, column_letter, start_row_index, end_row_index):

    ##-1 because start index with 0
    column_index = openpyxl.utils.column_index_from_string(column_letter) - 1  

        # Get cell values in the specified column for the given range of row indices
    column_values = []
    ##have to definition real row like 17 is Q
    for row in cbr_file.iter_rows(min_row=start_row_index+1, max_row=end_row_index-1, min_col=column_index+1, max_col=column_index+1, values_only=True):
        column_values.extend(row)

 
    return column_values





def find_row_index_with_keywords_in_column_b_view_name(cbr_file, keyword):

      # Find row index containing keyword
        row_index = None
        for idx, cell in enumerate(cbr_file['B'], start=1):
            if str(cell.value).lower() == keyword.lower():
                row_index = idx
                break

        return row_index


def search_keywords_right(worksheet, keywords):
    if worksheet is None:
        return
    
    result = {}
    for row in worksheet.iter_rows(values_only=True):
        for cell_value in row:
            if isinstance(cell_value, str) and cell_value in keywords:
                keyword = cell_value
                # Get the index of the keyword
                keyword_index = row.index(cell_value)
                # Check if the keyword is not the last element in the row
                if keyword_index < len(row) - 1:
                    # Get the value to the right of the keyword
                    value = row[keyword_index + 1]
                    if value is None:
                        result[keyword] = ""
                    else:
                        print(f"Found '{keyword}' Value: {value}")
                        result[keyword] = str(value).lower()  # Convert to lower case if needed
    return result

def search_keywords_job_name(worksheet, keywords):
    if worksheet is None:
        return None
    
    result = {}
    for row_index, row in enumerate(worksheet.iter_rows(values_only=True)):
        for cell_index, cell_value in enumerate(row):
            if isinstance(cell_value, str) and cell_value in keywords:
                keyword = cell_value
                # Get the index of the keyword
                keyword_index = cell_index
                # Check if the keyword is not the last element in the row
                if keyword_index < len(row) - 1:
                    # Get the value to the right of the keyword
                    value = row[keyword_index + 1]
                    if value is not None:
                        print(f"Found '{keyword}' Value: {value}")
                        result[keyword] = str(value).lower()  # Convert to lower case if needed
                        return result  # Exit the function after finding the first value
    
    return result




                    


def search_keywords_source_filename(worksheet, keywords):
    if worksheet is None:
        return
    
    result = {}
    for row in worksheet.iter_rows(values_only=True):
        for cell_value in row:
            if isinstance(cell_value, str) and cell_value in keywords:
                keyword = cell_value
                # Get the index of the keyword
                keyword_index = row.index(cell_value)
                # Check if the keyword is not the last element in the row
                if keyword_index < len(row) - 1:
                    # Get the value to the right of the keyword
                    value = row[keyword_index + 1]
                    if value is None:
                        result[keyword] = ""
                    else:
                        print(f"Found '{keyword}' Value: {value}")
                        result[keyword] = str(value)
                        return result
    return result


def search_keywords_encryped_key(worksheet, keywords):
    if worksheet is None:
        return
    
    result = {}
    for row in worksheet.iter_rows(values_only=True):
        for cell_value in row:
            if isinstance(cell_value, str) and cell_value in keywords:
                keyword = cell_value
                # Get the index of the keyword
                keyword_index = row.index(cell_value)
                # Check if the keyword is not the last element in the row
                if keyword_index < len(row) - 1:
                    # Get the value to the right of the keyword
                    value = row[keyword_index + 4]
                    if value is None:
                        result[keyword] = ""
                    else:
                        print(f"Found '{keyword}' Value: {value}")
                        result[keyword] = str(value).lower()  # Convert to lower case if needed
    return result

def search_keywords_sys_id(worksheet, keywords):
    if worksheet is None:
        return
    
    result = {}
    for row in worksheet.iter_rows(values_only=True):
        for i, cell_value in enumerate(row):
            if isinstance(cell_value, str) and cell_value in keywords:
                keyword = cell_value
                # Get the index of the keyword
                keyword_index = i
                # Check if the keyword is not the last element in the row
                if keyword_index < len(row) - 1:
                    # Get the value to the right of the keyword
                    value = row[keyword_index + 6]
                    if keyword not in result:
                        result[keyword] = []
                    if value is None:
                        result[keyword].append("")
                    else:
                        print(f"Found '{keyword}' Value: {value}")
                        result[keyword].append(str(value).lower())  # Convert to lower case if needed
    return result


def search_keywords_right_list(worksheet, keywords):
    if worksheet is None:
        return
    
    result = {}
    for row in worksheet.iter_rows(values_only=True):
        for i, cell_value in enumerate(row):
            if isinstance(cell_value, str) and cell_value in keywords:
                keyword = cell_value
                # Get the index of the keyword
                keyword_index = i
                # Check if the keyword is not the last element in the row
                if keyword_index < len(row) - 1:
                    # Get the value to the right of the keyword
                    value = row[keyword_index + 1]
                    if keyword not in result:
                        result[keyword] = []
                    if value is None:
                        result[keyword].append("")
                    else:
                        print(f"Found '{keyword}' Value: {value}")
                        result[keyword].append(str(value).lower())  # Convert to lower case if needed
    return result






def search_keywords_ingestion_method(worksheet, keywords):
    if worksheet is None:
        return

    result = {}
    for row_index, row in enumerate(worksheet.iter_rows(values_only=True)):
        for cell_index, cell_value in enumerate(row):
            if cell_value in keywords:
                keyword = cell_value
                value = worksheet.cell(row=row_index + 2, column=cell_index + 1).value
                if value == None:
                    result[keyword] = ""
                else:
                    print(f"Found '{keyword}' Value: {value}")
                    result[keyword] = value.lower()
                # If the keyword is found, break out of the loop
                if keyword == "Ingestion Method":
                    return result

    return result







def remove_quotes(cell_value):
    return cell_value.replace("'", "").replace('"', '')


def search_keywords_below(worksheet, keywords):
    if worksheet is None:
        return

    result = {}
    for row_index, row in enumerate(worksheet.iter_rows(values_only=True)):
        for cell_index, cell_value in enumerate(row):
            if cell_value in keywords:
                keyword = cell_value
                value = worksheet.cell(row=row_index + 2, column=cell_index + 1).value
                if value == None:
                    result[keyword] = ""
                else:
                    print(f"Found '{keyword}' Value: {value}")
                    result[keyword] = value.lower()

    return result

def search_keywords_left(worksheet, keywords):
    if worksheet is None:
        return
    
    result = {}
    for row in worksheet.iter_rows(values_only=True):
        for cell_value in row:
            if cell_value in keywords:
                keyword = cell_value
                value = row[row.index(cell_value) - 1]
                if value == None:
                    result[keyword] = ""
                else:
                    print(f"Found '{keyword}' Value: {value}")
                    result[keyword] = value.lower()
    return result

def remove_quotes(cell_value):
    return cell_value.replace("'", "").replace('"', '')


def get_column_params(worksheet, extraction_logic):
    # column M
    column_to_search = 13
    # column O
    column_to_get_value = 15
    
    keywords = extraction_logic.split(',')
    result_map = {}
    
    for keyword in keywords:
        try:
            for row in range(1, worksheet.max_row + 1):
                if worksheet.cell(row=row, column=column_to_search).value == keyword.strip():
                    value_in_column_O = worksheet.cell(row=row, column=column_to_get_value).value
                    result_map[keyword.strip()] = value_in_column_O
        except ValueError as e:
            result_map[keyword.strip()] = str(e)
                   
    return result_map



def column_params(extraction_logic, result_map):
    
    keywords = extraction_logic.split(',')
    
    p_dd = "{{ptn_dd}}" if 'ptn_dd' in extraction_logic else ""
    p_mm = "{{ptn_mm}}" if 'ptn_mm' in extraction_logic else ""
    p_qtr = "{{ptn_qtr}}" if 'ptn_qtr' in extraction_logic else ""
    p_yyyy = "{{ptn_yyyy}}" if 'ptn_yyyy' in extraction_logic else ""

    dd = result_map.get("ptn_dd", "").lower()
    mm = result_map.get("ptn_mm", "").lower()
    qtr = result_map.get("ptn_qtr", "").lower()
    yyyy = result_map.get("ptn_yyyy", "").lower()

    keyword_to_value = {
        "ptn_dd": f"cast('{p_dd}' as {dd})",
        "ptn_mm": f"cast('{p_mm}' as {mm})",
        "ptn_qtr": f"cast('{p_qtr}' as {qtr})",
        "ptn_yyyy": f"cast('{p_yyyy}' as {yyyy})"
    }

    # Use a list comprehension to extract the corresponding values
    result = {keyword:keyword_to_value[keyword] for keyword in keywords if keyword in keyword_to_value}

    # Join the values into a comma-separated string
    # result = ',\n'.join(values)
    # result = result.replace(", ", ",\n")
    # print('column param', result)

    return result

def replace_stars(worksheet):
    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
        for cell in row:
            if cell.value is not None and isinstance(cell.value, str):
                cell.value = cell.value.replace('*', '')


             




def convert_format_date(sourceFileName):
    # Split the path by '/'

    if 'ddmmyy' in sourceFileName:
        # Split the filename at the pattern
      # Split the filename at the pattern
        split_index = sourceFileName.index('ddmmyy')
        prefix = sourceFileName[:split_index]
        date_part = sourceFileName[split_index:]
           
        
        # Convert 'ddmmyy' to uppercase
        date_part = date_part.replace('ddmmyy', 'DDMMYY')
        print(date_part)

        # Define regular expressions for different date patterns
        patterns = {
            r'MM': r'{{ptn_mm}}',
            r'DD': r'{{ptn_dd}}',
            r'YY' :r'{{ptn_yyyy[2:]}}'
        }  

        for pattern, replacement in patterns.items():
             date_part = re.sub(pattern, replacement, date_part)

            # If there was a prefix, join it back with the modified date part
        if prefix:
            modified_sourceFileName = prefix + date_part

            print(modified_sourceFileName)
            print(modified_sourceFileName)
            print(modified_sourceFileName)
        
        return modified_sourceFileName
    

    elif '%Y%m%d' in sourceFileName:

        split_index = sourceFileName.index('%Y%m%d')
        prefix = sourceFileName[:split_index]
        date_part = sourceFileName[split_index:]


         # Convert 'ddmmyy' to uppercase
        date_part = date_part.replace('%Y%m%d', 'YYYYMMDD')
        print(date_part)

        # Define regular expressions for different date patterns
          # Define regular expressions for different date patterns
        
        patterns = {
            r'YYYY': r'{{ptn_yyyy}}',
            r'MM': r'{{ptn_mm}}',
            r'DD': r'{{ptn_dd}}',
            # r'yyyy': r'{{ptn_yyyy}}',
            # r'mm': r'{{ptn_mm}}',
            # r'dd': r'{{ptn_dd}}'
        }

        for pattern, replacement in patterns.items():
             date_part = re.sub(pattern, replacement, date_part)

            # If there was a prefix, join it back with the modified date part
        if prefix:
            modified_sourceFileName = prefix + date_part

            print(modified_sourceFileName)
            print(modified_sourceFileName)
            print(modified_sourceFileName)
        
        return modified_sourceFileName
    
    elif '%%TRANDATE' in sourceFileName:

        split_index = sourceFileName.index('%%TRANDATE')
        prefix = sourceFileName[:split_index]
        date_part = sourceFileName[split_index:]


         # Convert 'ddmmyy' to uppercase
        date_part = date_part.replace('%%TRANDATE', 'DDMMYY')
        print(date_part)

        # Define regular expressions for different date patterns
        patterns = {
            r'MM': r'{{ptn_mm}}',
            r'DD': r'{{ptn_dd}}',
            r'YY' :r'{{ptn_yyyy[2:]}}'
        }  

        for pattern, replacement in patterns.items():
             date_part = re.sub(pattern, replacement, date_part)

            # If there was a prefix, join it back with the modified date part
        if prefix:
            modified_sourceFileName = prefix + date_part

            print(modified_sourceFileName)
            print(modified_sourceFileName)
            print(modified_sourceFileName)
        
        return modified_sourceFileName

    
    else :
        path_parts = sourceFileName.split('/')
        # Get the last part which contains the filename
        filename = path_parts[-1]
        # Split the filename by '_'
        filename_parts = filename.split('_')
        # Get the last part which contains the date
        date_part = filename_parts[-1]

        if 'yyyymmdd' in date_part.lower():
            date_part = date_part.replace('yyyymmdd', 'YYYYMMDD')
        if 'yyyyMMdd' in date_part:
            date_part = date_part.replace('yyyyMMdd', 'YYYYMMDD')
        if 'yyyy-mm-dd' in date_part.lower():
            date_part = date_part.replace('yyyy-mm-dd', 'YYYY-MM-DD')

    

        # Define regular expressions for different date patterns
        patterns = {
            r'YYYY': r'{{ptn_yyyy}}',
            r'MM': r'{{ptn_mm}}',
            r'DD': r'{{ptn_dd}}',
            # r'yyyy': r'{{ptn_yyyy}}',
            # r'mm': r'{{ptn_mm}}',
            # r'dd': r'{{ptn_dd}}'
        }

        # Replace date patterns in the date part of the filename
        for pattern, replacement in patterns.items():
            date_part = re.sub(pattern, replacement, date_part)


        
        # Replace the original date part with the modified one
        filename_parts[-1] = date_part
        
        # Join the filename parts back together
        filename = '_'.join(filename_parts)
        
        # Replace the original filename with the modified one in the path
        path_parts[-1] = filename
        
        # Join the path parts back together
        modified_sourceFileName = '/'.join(path_parts)
    
    return modified_sourceFileName






def transform_file_extension(file_name, new_format):

     
    base_name,extension  = os.path.splitext(file_name)
    new_file_name = f"{base_name}.{new_format}"
    return new_file_name





def extract_number_with_regex(input_string):
    # Define a regex pattern to match numbers
    pattern = r"\d+"
    # Find all matches of the pattern in the input string
    matches = re.findall(pattern, input_string)
    # Get the first match (assuming there's only one number)
    number = int(matches[0]) if matches else None
    return number

    

def fixlength_and_concatenate(string):
    # Split the string using underscore
    parts = string.split('_')

    parts.pop(0)
    # ingest_spec_cardlink_cpbcrd_fixed_length.json

    # Join the parts and concatenate with "extrct_"
    concatenated_word = "fixlength_" + '_'.join(parts)
    
    return concatenated_word



def get_all_value_from_column(worksheet,column_letter):
    column_values = worksheet[column_letter]
    result = []
    for i, cell in enumerate(column_values, start=1):
        result.append((i, cell.value))
    return result

def get_all_value_from_column_check_null(worksheet, column_letter):
    column_values = worksheet[column_letter]
    result = []
    for i, cell in enumerate(column_values, start=1):
        value = cell.value
        if value is None:
            value = "NULL"
        result.append((i, value))
    return result




def find_word_position_in_excel(sheet, word):
    # Iterate through each cell in the specified sheet
    for row in sheet.iter_rows():
        for cell in row:
            # Check if the cell contains the word
            if isinstance(cell.value, str) and word in cell.value:
                # Find the position of the word within the cell value
                position = cell.value.find(word)
                # Return the position and cell address
                return position, cell.coordinate

    # If the word is not found in any cell, return None
    return None


def retrieve_values_from_column(sheet, start_row, end_row):
    retrieved_values = []
    # Iterate through each cell in the specified range (column B)
    for row in range(start_row, end_row + 1):
    
            cell_W = sheet[f'W{row}']
            retrieved_values.append(cell_W.value)
    return retrieved_values


def extract_row_number(cell_reference):
    # Split the cell reference into column and row parts
    column, row = cell_reference[:1], cell_reference[1:]
    # Extract the row number and return it
    row_number = int(row)
    return row_number



def format_string_for_diff_field(value1, value2):
    return '"{}":"cast(`{}` as string)"'.format(value1, value2)

